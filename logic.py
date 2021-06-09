import datetime
import getpass
import json
import os
import queue
import socket
import sys
import threading
import time
from multiprocessing import Process
from binance import AsyncClient, BinanceSocketManager

import gnupg
import mysql.connector
import pandas
import requests
import yaml
from colorama import Fore
from colorama import Style
from openpyxl import load_workbook

import banking
import data_visual
from hype import Hype
from trade import Operation

_print_list = []

async def arbo():
    d = {}
    with open("config.yaml") as f:
        d = yaml.safe_load(f)
    ip_mon = d["dip"]
    gpg = gnupg.GPG(d["gpg"][:-1])
    _list = []
    _trade_list = []
    exchange_list = []
    print(f"""{Fore.RED}                                                                                          
                                                                                              
                                                                   .%%%#                      
                                                               *#%%%%%%%%,                    
                                                           /#%%%%%%%%%%%%%%                   
                                                       %%%%%%%%%%%%%%%%%%%%%                  
                                                   #%%%%%%%%%%%%%%%%%%%%%%%%%%                
                                                  %%%%%%%%%%%%%%%%%%%%%%%%%%%%%               
                                                  %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#             
                                      /%/        %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%#            
                                  %#%%%%%%      (%%%%%%%%%%%%(     *#%%%%%%%%%%%%%%           
                              %%%%%%%%%%%%%%    %%%%%%%#%.           #%%%%%%%%%%%%            
                         ,#%%%%%%%%%%%%%%%%%%  %%%%%%                 (%%%%%%%%%%             
                     (%%%%%%%%%%%%%%%%%%%%%%%%%#/                       %%%%%%%%*             
                    %%%%%%%%%%%%%%%%%%%%%%%%%%%%                         #%%%%%#              
                   #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%                         %%%#               
                   %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%                         %%                
                  %%%%%%%%%%%%%#    .%%%%%%%%%%%%%%%*                     .%#                 
                 .%%%%%%%%#,          %%%%%%%%%%%%%  #                /%*                     
                 %%%%%%                *%%%%%%%%%%(   %.          %%                          
                (%#                      #%%%%%%%%     .%    .##                              
                ,%                        (#%%%%#        ##/                                  
                  #                         %%%%                                              
                   ((                        #%                                               
                     #                      #%                                                
                      #*                #%                                                    
                        %          *%(                                                        
                         %.    #%.                                                            
                           %%                                                                 
                                                                                              
      {Style.RESET_ALL}""")
    ok_pass = False
    while not ok_pass:
        try:
            passphrase = getpass.getpass("[?] Please provide master password to continue:")
            with open("telegram.gpg", "rb") as tg_f:
                status_tg = gpg.decrypt_file(file=tg_f, passphrase=passphrase)
            with open("keydict.gpg", "rb") as kd_f:
                status_kd = gpg.decrypt_file(file=kd_f, passphrase=passphrase)
            with open("dbinfo.gpg", "rb") as db_f:
                status_db = gpg.decrypt_file(file=db_f, passphrase=passphrase)
            printin_data = json.loads(str(status_kd).strip().replace("\n", ""))
            tg_data = json.loads(str(status_tg).strip().replace("\n", ""))
            db_data = json.loads(str(status_db).strip().replace("\n", ""))
            ok_pass = True
        except json.decoder.JSONDecodeError as err:
            print(f"{Fore.RED}[!] Wrong password!{Style.RESET_ALL}")
            log("ERR", err)
            pass
    if d['balancing']['banking']:
        hype = Hype()
        hype.login(d['balancing']['username'], getpass.getpass('[?] Please provide banking password'),
                   getpass.getpass('[?] Please provide DOB'))
        OTP = getpass.getpass('[?] Please provide hype OTP to continue:')
        hype.otp2fa(OTP)
        try:
            hype.get_balance()
        except banking.Banking.AuthenticationFailure:
            exit(75)
    try:
        trt_apikey = str(printin_data["trt_apikey"])
        trt_secret = str(printin_data["trt_secret"])
        exchange_list.append("trt")
    except:
        exit(1)
    try:
        krk_apikey = str(printin_data["krk_apikey"])
        krk_secret = str(printin_data["krk_secret"])
        exchange_list.append("krk")
    except:
        krk_apikey = ""
        krk_secret = ""
    try:
        bnb_apikey = str(printin_data["bnb_apikey"])
        bnb_secret = str(printin_data["bnb_secret"])
        exchange_list.append("bnb")
    except:
        exit(3)
    time_list = []
    print(f"{Fore.GREEN}\nThank you{Style.RESET_ALL}")

    while not check_api_connection():
        time.sleep(1)
    op = Operation(trt_apikey, trt_secret, krk_apikey, krk_secret, bnb_apikey, bnb_secret, exchange_list)
    op.threadCreation()
    time.sleep(2)
    taker_fee, maker_fee = dict(), dict()
    taker_fee['trt'] = float(d["taker_fee_trt"])
    taker_fee['bnb'] = float(d["taker_fee_bnb"])
    maker_fee['trt'] = float(d["taker_fee_trt"])
    maker_fee['bnb'] = float(d["taker_fee_bnb"])
    checkbalance = True
    _tglist = []
    if d["graph"]:
        g = Process(target=data_visual.ru)
        g.start()
    f = open("version", "r")
    ver = f.read()
    all_balance = op.balancethreading()
    s = socket.socket()
    print(str(d["dip"]))
    dir_path = os.path.dirname(os.path.realpath(__file__))
    last_h = 0
    try:
        s.settimeout(10)
        s.connect((str(d["dip"]).rstrip("\n"), 30630))
    except ConnectionRefusedError as err:
        log("ERR", err)
        pass
    except TimeoutError as err:
        log("ERR", err)
        pass
    except socket.timeout as err:
        log("ERR", err)
        pass
    q_act = queue.Queue()
    t_action = threading.Thread(target=getaction, args=(q_act,))
    t_action.start()
    client = await AsyncClient.create()
    bm = BinanceSocketManager(client)
    ds = bm.depth_socket('BTCEUR', depth=BinanceSocketManager.WEBSOCKET_DEPTH_5, interval=100)
    already_saved = False
    only_see = bool(int(d["only_see"]))
    this_sec = str(int(time.time()))[-4:]
    count = 0
    actual = 0
    while True:
        _start_time = time.time()
        _query_time = time.time()
        eff = 0
        prod = 0
        try:
            if checkbalance:
                print(f"{Fore.YELLOW}[#] RETRIEVING BALANCE{Style.RESET_ALL}")
                all_balance = op.balancethreading()
                checkbalance = False
                time.sleep(int(d['sleep_balance']))
            price_dict = op.querythread()
            #requests_used=price_dict['bnb'].headers['x-mbx-used-weight-1m']
            _query_time = time.time() - _query_time
            async with ds as tscm:
                res = await tscm.recv()
                print(res['asks'][0])
            try:
                asks_data_bnb = price_dict["bnb"]['asks'][0]
                bids_data_bnb = price_dict["bnb"]['bids'][0]
                asks_data_trt = price_dict["trt"]['asks'][0]
                bids_data_trt = price_dict["trt"]['bids'][0]
            except TypeError as err:
                print(f"{Fore.RED}[#] ERROR WHILE FETCHING DATA [typeError - nonetype]{Style.RESET_ALL}")
                log("ERR", err)
                continue
            asks = bids = {}
            asks['bnb'] = round(float(price_dict["bnb"]['asks'][0][0]), 2)
            bids['bnb'] = round(float(price_dict["bnb"]['bids'][0][0]), 2)
            asks['trt'] = round(float(price_dict["trt"]['asks'][0]['price']), 2)
            bids['trt'] = round(float(price_dict["trt"]['bids'][0]['price']), 2)
            last_bid = 0
            last_ask = 0
            depth = 0
            bal_list = False
            if not q_act.empty():
                action = last(q_act)
                if action == "STOP" and not only_see:
                    only_see = True
                elif action == "GO" and only_see:
                    only_see = False
            os.system('cls' if os.name == 'nt' else 'clear')
            a = datetime.datetime.now()
            if not str(a.microsecond)[:-5]:
                small_index = 0
            else:
                small_index = str(a.microsecond)[:-5]
            _trade_list.clear()
            balance_score = dict()
            balance_score["eur"] = (all_balance["trteur"] - all_balance["bnbeur"]) / (
                    all_balance["trteur"] + all_balance["bnbeur"])
            balance_score["btc"] = (all_balance["trtbtc"] - all_balance["bnbbtc"]) / (
                    all_balance["trtbtc"] + all_balance["bnbbtc"])
            print(f"{Fore.MAGENTA}[!] ARBOBOTTI VERSION %s, MURINEDDU CAPITAL 2021{Style.RESET_ALL}\n" % (ver))
            print(
                f"{Fore.LIGHTCYAN_EX}[i] %s{Style.RESET_ALL}\tINDEX: {Fore.LIGHTCYAN_EX}%s - %s{Style.RESET_ALL}\tTHREAD_POOL:{Fore.LIGHTCYAN_EX} %s{Style.RESET_ALL}\tONLY_SEE: {Fore.LIGHTCYAN_EX} %d{Style.RESET_ALL}\tPERF: {Fore.LIGHTCYAN_EX} %d{Style.RESET_ALL} cycles/s\tREQUESTS: {Fore.LIGHTCYAN_EX} %d{Style.RESET_ALL}/1200" % (
                    a.strftime("%d/%m/%Y %H:%M:%S"), str(int(time.time()))[-4:], small_index, str(op.len), only_see,
                    actual,int(op.get_requests_used())))
            next_one = False
            for exchange_a in exchange_list:
                for exchange_b in exchange_list:
                    if exchange_a == exchange_b:
                        next_one = True
                        continue
                    if next_one:
                        # info(exchange_a, exchange_b, all_balance, asks, bids, taker_fee, maker_fee=None)
                        info(exchange_a, exchange_b, all_balance, asks, bids, taker_fee)
                next_one = False
            # info('trt', 'bnb', all_balance, asks, bids, taker_fee)

            print("[i] FETCHED TAKER FEE       %s: %.4f%%;      %s: %.4f%%" % (
                exchange_list[0].upper(), taker_fee['trt'] * 100, exchange_list[1].upper(), taker_fee['bnb'] * 100))
            print("[i] FETCHED MAKER FEE       %s: %.4f%%;      %s: %.4f%%" % (
                exchange_list[0].upper(), maker_fee['trt'] * 100, exchange_list[1].upper(), maker_fee['bnb'] * 100))
            print("[i] BALANCE SCORE: %.4f" % (
                balance_score["eur"]))
            if (bids['trt'] * (1 - taker_fee['trt'])) - (asks['bnb'] * (1 + taker_fee['bnb'])) > int(d["threshold"]):
                low_balance = False
                print(f"{Fore.CYAN}[#] %.2f < %.2f BUY %s | SELL TRT DIFF: %.2f (MENO FEE): %.3f" % (
                    asks['bnb'], bids['trt'], exchange_list[1].upper(), bids['trt'] - asks['bnb'],
                    (bids['trt'] * (1 + taker_fee['trt'])) - (asks['bnb'] * (1 + taker_fee['bnb']))))
                depth = float(min(bids_data_trt['amount'],
                                  float(asks_data_bnb[1])))
                balance = min(all_balance["trtbtc"], all_balance["bnbeur"] / asks['bnb'])
                if balance < depth:
                    depth = round(float(balance * float(d["max_each_trade"])), 4)
                    print(f"{Fore.MAGENTA}[#] PARTIAL FILLING, BALANCE LOWER THAN DEPTH{Style.RESET_ALL}")
                    print(f"{Fore.MAGENTA}[#] DEPTH %f{Style.RESET_ALL}" % (depth))

                    if depth == 0:
                        print(f"{Fore.RED}[#] BALANCE IS LOW, PLEASE DEPOSIT TO CONTINUE{Style.RESET_ALL}")
                        low_balance = True
                else:
                    print(f"{Fore.GREEN}[#] COMPLETE FILLING{Style.RESET_ALL}")
                    depth = round(float(depth * float(d["max_each_trade"])), 4)
                    print(f"{Fore.GREEN}[#] DEPTH %f{Style.RESET_ALL}" % depth)

                if not low_balance and depth > float(d["min_balance"]):
                    print(f"{Fore.CYAN}[#] DEPTH %f BTC" % depth)
                    eff = (depth * bids['trt'] * (1 - taker_fee['trt'])) - (
                            depth * asks['bnb'] * (1 + taker_fee['bnb']))
                    prod = eff / (depth * bids['trt'])
                    print("[#] GAIN DOPO FEE EFF %f € | PROD %f ¢/€" % (eff, prod * 100))
                    print(f"[#] NEED %.3f EUR | %f BTC{Style.RESET_ALL}" % (asks['bnb'] * depth, depth))
                    last_ask = asks['bnb']
                    last_bid = bids['trt']
                    if prod * 100 > float(d["prod_threshold"]) and not only_see:
                        checkbalance = True
                        print(f"{Fore.GREEN}[#] TRADE{Style.RESET_ALL}")
                        print(f"{Fore.YELLOW}[H] SELL %f BTC ON TRT, BUYING %f (%f) BTC ON BNB{Style.RESET_ALL}" % (
                            depth, depth, depth * last_ask))
                        resp_dict = op.tradethreading("sell", "trt", "BTCEUR", depth, last_bid, "buy", "bnb", "BTCEUR",
                                                      depth,
                                                      last_ask)
                        print("BNB", resp_dict["bnb"], "\nTRT", resp_dict["trt"])
                        log("TRADE", "BNB" + str(resp_dict["bnb"]) + "\nTRT" + str(resp_dict["trt"]))
                        try:
                            status = (resp_dict["bnb"]["status"], resp_dict["trt"]["status"])
                        except KeyError as err:
                            print(f"{Fore.RED}[!] ERROR RETRIEVING STATUS{Style.RESET_ALL}")
                            status = (resp_dict["bnb"]["status"], resp_dict["trt"]["errors"][0]["message"])
                            log("ERR", str(resp_dict["bnb"]["status"]) + str(resp_dict["trt"]["errors"][0]["message"]))
                            log("ERR", err)
                        if status[0] == "ERROR" or status[1] == "ERROR":
                            print(f"{Fore.RED}[$] TRADE ERROR MSG: [%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"][0].upper(), resp_dict["bnb"]))
                            log("ERROR", str(resp_dict["trt"][0].upper()) + str(resp_dict["bnb"]))
                            time.sleep(0)
                        else:
                            print(f"{Fore.GREEN}[#] SOUNDS GOOD! ORDER STATUS:[%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"]["status"].upper(), resp_dict["bnb"]["status"]))
                            bal_list = True
                            time.sleep(int(d["sleep_check_order"]))
                            _trade_list.append(
                                [datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "BUY", exchange_list[1].upper(),
                                 str(depth).replace(",", "."),
                                 str(last_ask).replace(".", ","),
                                 "SELL", "TRT", str(last_bid).replace(".", ","), float(all_balance["bnbbtc"]),
                                 float(all_balance["trtbtc"]),
                                 float(all_balance["bnbeur"]), float(all_balance["trteur"]),
                                 float(all_balance["trteur"] + all_balance["bnbeur"] + (
                                         all_balance["bnbbtc"] + all_balance["trtbtc"]) * last_ask)])
                            pass
                            # EXECUTED OR SUCCESS
                            # IF BOTH ORDER ARE NOT COMPLETED, DELETE ORDER
                else:
                    print(f"{Fore.RED}[$] TOO LOW BALANCE, PLEASE DEPOSIT{Style.RESET_ALL}")
                    checkbalance = True
            elif (bids['bnb'] * (1 - taker_fee['bnb'])) - (asks['trt'] * (1 + taker_fee['trt'])) > int(d["threshold"]):
                low_balance = False
                depth = float(min(asks_data_trt['amount'],
                                  float(bids_data_bnb[1])))
                balance = min(all_balance["bnbbtc"], all_balance["trteur"] / asks['trt'])
                print(
                    f"{Fore.CYAN}[!] %.2f < %.2f BUY TRT | SELL %s DIFF: %.2f (MENO FEE): %.3f | DEPTH: %.8f | MINBAL: %.8f{Style.RESET_ALL}" % (
                        asks['trt'], bids['bnb'], exchange_list[1].upper(), bids['bnb'] - asks['trt'],
                        (bids['bnb'] * (1 + taker_fee['bnb'])) - (asks['trt'] * (1 + taker_fee['trt'])), depth,
                        balance))
                if balance < depth:
                    depth = round(float(balance * float(d["max_each_trade"])), 4)
                    print(f"{Fore.MAGENTA}[#] PARTIAL FILLING, BALANCE LOWER THAN DEPTH{Style.RESET_ALL}")
                    print(f"{Fore.MAGENTA}[#] DEPTH %f{Style.RESET_ALL}" % (depth))
                    if depth == 0:
                        print(f"{Fore.MAGENTA}[#] BALANCE IS LOW, PLEASE DEPOSIT TO CONTINUE{Style.RESET_ALL}")
                        low_balance = True
                else:
                    print(f"{Fore.GREEN}[#] COMPLETE FILLING{Style.RESET_ALL}")
                    depth = round(float(depth * float(d["max_each_trade"])), 4)
                    print(f"{Fore.GREEN}[#] DEPTH %f{Style.RESET_ALL}" % (depth))

                if not low_balance and (depth > float(d["min_balance"])):
                    print(f"{Fore.CYAN}[!] DEPTH %f BTC" % depth)
                    eff = (depth * bids['bnb'] * (1 - taker_fee['bnb'])) - (
                            depth * asks['trt'] * (1 + taker_fee['trt']))
                    prod = eff / (depth * bids['bnb'])
                    print("[i] GAIN DOPO FEE EFF %f € | PROD %f ¢/€" % (eff, prod * 100))
                    print(f"[i] NEED %.3f EUR | %f BTC{Style.RESET_ALL}" % (asks['trt'] * depth, depth))
                    last_ask = asks['trt']
                    last_bid = bids['bnb']
                    if prod * 100 > float(d["prod_threshold"]) and not only_see:
                        checkbalance = True
                        print(f"{Fore.GREEN}[#] TRADE{Style.RESET_ALL}")
                        print(f"{Fore.YELLOW}[H] SELL %f BTC ON BNB, BUYING %f (%f) BTC ON TRT{Style.RESET_ALL}" % (
                            depth, depth, depth * last_ask))
                        resp_dict = op.tradethreading("buy", "trt", "BTCEUR", depth, last_ask,
                                                      "sell", exchange_list[1], "BTCEUR", depth, last_bid)
                        print("BNB", resp_dict["bnb"], "\nTRT", resp_dict["trt"])
                        log("TRADE", "BNB" + str(resp_dict["bnb"]) + "\nTRT" + str(resp_dict["trt"]))
                        try:
                            status = (resp_dict["bnb"]["status"], resp_dict["trt"]["status"])
                        except KeyError as err:
                            print(f"{Fore.RED}[!] ERROR RETRIEVING STATUS{Style.RESET_ALL}")
                            status = (resp_dict["bnb"]["status"], resp_dict["trt"]["errors"][0]["message"])
                            log("ERR", str(resp_dict["bnb"]["status"]) + str(resp_dict["trt"]["errors"][0]["message"]))
                            log("ERR", err)
                        if status[0] == "ERROR" or status[1] == "ERROR":
                            print(f"{Fore.RED}[$] TRADE ERROR MSG: [%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"][0].upper(), resp_dict["bnb"]))
                            log("ERROR", str(resp_dict["trt"][0].upper()) + (resp_dict["bnb"]))
                            time.sleep(0)
                            pass
                            # TODO: IMPROVE ERROR CHECK
                        else:
                            print(f"{Fore.GREEN}[#] SOUNDS GOOD! ORDER NO:[%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"]["status"].upper(), resp_dict["bnb"]["status"]))
                            bal_list = True
                            time.sleep(int(d["sleep_check_order"]))
                            _trade_list.append(
                                [datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "BUY", "TRT",
                                 str(depth).replace(".", ","), str(last_ask).replace(".", ","),
                                 "SELL", exchange_list[1].upper(), str(last_bid).replace(".", ","),
                                 float(all_balance["bnbbtc"]),
                                 float(all_balance["trtbtc"]),
                                 float(all_balance["bnbeur"]), float(all_balance["trteur"]),
                                 float(all_balance["trteur"] + all_balance["bnbeur"] + (
                                         all_balance["bnbbtc"] + all_balance["trtbtc"]) * last_ask), resp_dict["bnb"],
                                 resp_dict["trt"]])
                else:
                    print(f"{Fore.RED}[$] TOO LOW BALANCE, PLEASE DEPOSIT{Style.RESET_ALL}")
                    checkbalance = True

            # TODO:GROUP THE TWO CHECK FUNCTION TOGETHER (SIMPLER CODE AND UPDATE)
            _end_time = time.time()
            totaltime = _end_time - _start_time
            time_list.append(int(totaltime * 1000))
            _list.append([datetime.datetime.now(), asks['bnb'], bids['bnb'], asks['trt'], bids['trt'],
                          all_balance["bnbbtc"], all_balance["trtbtc"], all_balance["bnbeur"],
                          all_balance["trteur"], eff, prod, int(totaltime * 1000), round(bids['trt'] - asks['bnb'], 2),
                          round((bids['trt'] * (1 - taker_fee['trt'])) - (asks['bnb'] * (1 + taker_fee['bnb'])), 2),
                          round(bids['bnb'] - asks['trt'], 2),
                          round((bids['bnb'] * (1 - taker_fee['bnb'])) - (asks['trt'] * (1 + taker_fee['trt'])), 2)])
            if int(_end_time % int(d["save_interval"])) == 0 and not already_saved:
                print(f"{Fore.YELLOW}[!] SAVING...{Style.RESET_ALL}")
                if _list:
                    save_data_thread = threading.Thread(target=save_data, args=(_list, d["sep"],))
                    save_data_thread.start()
                    already_saved = True
                    try:
                        data = arbomonitor(s, only_see, last_h, d["name"])
                        if data == "go":
                            only_see = True
                        if "command" in data:
                            os.system("cd " + dir_path)
                            os.system("cd " + dir_path + " && " + data.split(":")[1])
                    except Exception as errr:
                        print(errr)
                        log("ERR", errr)
                        pass
                        try:
                            s.close()
                            s.connect((ip_mon, 30630))
                        except ConnectionRefusedError as err:
                            print(err)
                            log("ERR", err)
                            pass
                        except socket.gaierror as err:
                            print(err)
                            log("ERR", err)
                            pass
                        except Exception as err:
                            print(err)
                            log("ERR", err)
                            pass
                        except TimeoutError as err:
                            print(err)
                            log("ERR", err)
                            pass
            if int(str(int(_end_time))[-1]) > 1:
                already_saved = False
            if int(_end_time % int(d["balance_interval"])) == 0:
                checkbalance = True
            if _trade_list:
                all_balance = op.balancethreading()
                time.sleep(int(d['sleep_balance']))
                _trade_list.append(
                    ["", "", "", "", "", "", "", "", float(all_balance["bnbbtc"]), float(all_balance["trtbtc"]),
                     float(all_balance["bnbeur"]),
                     float(all_balance["trteur"]),
                     ((all_balance["bnbbtc"] + all_balance["trtbtc"]) * last_ask) +
                     float(all_balance["bnbeur"]) +
                     float(all_balance["trteur"])])
                print(f"{Fore.YELLOW}[!] SAVING TRADE LIST...{Style.RESET_ALL}")
                save_trade_thread = threading.Thread(target=save_trade,
                                                     args=(_trade_list, d["sep"], db_data, tg_data, d['name']))
                save_trade_thread.start()
                save_trade_thread.join()
                _trade_list.clear()
            if str(int(time.time()))[-4:] == this_sec:
                count += 1
            else:
                actual = count
                this_sec = str(int(time.time()))[-4:]
                count = 0
            last_h = sum(time_list[-100:]) / min(100, len(time_list))
            print(
                f"[-] ------------------------------------------------- {Fore.YELLOW}%d ms{Style.RESET_ALL} (%d ms(q) + %d ms(p)) - avg last %d ({Fore.YELLOW}%d ms{Style.RESET_ALL}) - global avg ({Fore.YELLOW}%d ms{Style.RESET_ALL})" % (
                    int(totaltime * 1000), int(_query_time * 1000),
                    (int(totaltime * 1000) - int(_query_time * 1000)), min(100, len(time_list)),
                    sum(time_list[-100:]) / min(100, len(time_list)), sum(time_list) / len(time_list)))
        except KeyboardInterrupt as err:
            log("ERR", err)
            sys.exit()


def save_data(_list, sep):
    df = pandas.DataFrame(_list)
    try:
        df.to_csv('filev2.csv', index=False, sep=',', mode='a', header=False, decimal=',')
        # append(df, filename='filev2.xlsx', startrow=None, sheet_name='Sheet1', truncate_sheet=True,engine="openpyxl")
        _list.clear()
    except FileNotFoundError as err:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATAprint [file not found]{Style.RESET_ALL}")
        log("ERR", err)
    except PermissionError as err:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATAprint [resource busy]{Style.RESET_ALL}")
        log("ERR", err)
    except:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATAprint [generic error]{Style.RESET_ALL}")
        log("ERR", "generic except")


def save_trade(_list, sep, db_data, tg_data, whoami):
    print(_list)
    df = pandas.DataFrame(_list)
    try:
        # with open('file_trade.xlsx', 'a') as f:
        df.to_csv("file_trade.csv", sep=',', mode='a', index=False, header=False, decimal=',')
        # append(df, filename='file_trade.xlsx', startrow=None, sheet_name='Sheet1', truncate_sheet=True,engine="openpyxl")
    except FileNotFoundError as err:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO TRADELIST [file not found]{Style.RESET_ALL}")
        log("ERR", err)
    except PermissionError as err:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO TRADELIST [resource busy]{Style.RESET_ALL}")
        log("ERR", err)
    except TypeError as err:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO TRADELIST [type error]{Style.RESET_ALL}")
        log("ERR", err)
    except:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATAprint [generic error]{Style.RESET_ALL}")
        log("ERR", "generic except")
    try:
        telegram(_list, tg_data, whoami)
    except Exception as err:
        print(f"{Fore.RED}[ERR] ERRORE INVIO TELEGRAM [generic error]{Style.RESET_ALL}")
        log("ERR", err)
        print(err)
        exit(100)
    try:
        db(_list, db_data)
    except Exception as err:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATABASE [generic error]{Style.RESET_ALL}")
        log("ERR", err)
        print(err)
        exit(100)
    pass


def log(log_type, message, thread=None):
    if thread:
        if log_type == "ERR":
            with open("errorlog", "a") as f:
                f.write("[" + log_type + "] " + datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S") + " " + str(
                    message) + "\nEND\n")
            exit(20)
        if log_type == 'TRADELIST' or log_type == 'TRADE':
            with open("tradelog", "a") as f:
                f.write("[" + log_type + "] " + datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S") + " " + str(
                    message) + "\nEND\n")
            exit(20)
    else:
        thread_l = threading.Thread(target=log, args=(log_type, message, 1))
        thread_l.start()
        return 0


def append(df, filename, startrow=None, sheet_name='Sheet1', truncate_sheet=True, engine="xlrd"):
    writer = pandas.ExcelWriter(filename, engine=engine)
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError as err:
        print("e diocan pero")
        log("ERR", err)
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow)
    writer.save()


def db(_list, db_data):
    server = db_data["dbhost"]
    database = db_data["dbname"]
    username = db_data["dbuser"]
    password = db_data["dbpass"]
    port = int(db_data["dbport"])
    conn = mysql.connector.connect(host=server, user=username, password=password, port=port, database=database)
    cursor = conn.cursor()
    date = datetime.datetime.strptime(_list[0][0], "%d/%m/%Y %H:%M:%S").strftime('%Y-%m-%d %H:%M:%S')
    status = "err" in _list[len(_list)].lower() or "err" in _list[len(_list) - 1].lower()
    add_trade = (
        "INSERT INTO `tradelist` "
        "(`side1`,`exch1`,`ask`,`side2`,`exch2`,"
        "`bid`,`depth`,"
        "`bnbbtc_in`,`trtbtc_in`,`bnbeur_in`,`trteur_in`,`bal_in`,"
        "`bnbbtc_en`,`trtbtc_en`,`bnbeur_en`,`trteur_en`,`bal_en`,`gain`,`date`,"
        "`order_id_1`,`order_id_2`,`success`) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);")
    data_trade = (
        _list[0][1], _list[0][2], _list[0][4].replace(",", "."), _list[0][5], _list[0][6],
        _list[0][7].replace(",", "."), _list[0][3].replace(",", "."),
        _list[0][8], _list[0][9], _list[0][10], _list[0][11], _list[0][12],
        _list[1][8], _list[1][9], _list[1][10], _list[1][11], _list[1][12],
        str(round(float(_list[1][11]) + float(_list[1][10]) - float(
            _list[0][11]) - float(_list[0][10]), 5)), date,
        _list[0][len(_list) - 1], _list[0][len(_list) - 2], str(status))

    cursor.execute(add_trade, data_trade)
    conn.commit()
    cursor.close()
    conn.close()


def telegram(_list, tg_data, whoami):
    log(log_type="TRADELIST", message=str(_list))
    message = ("EXECUTED TRADE AT " + str(_list[0][0]) + ":\n - BY " + whoami + "\nBOUGHT <b>" + str(
        round(float(_list[0][3].replace(",", ".")), 8)) + "</b> $BTC <b>" + str(
        _list[0][4]) + "</b> ON <code>" + str(
        _list[0][2]) + "</code> SOLD <b>" + str(_list[0][7]) + "</b> ON <code>" + str(
        _list[0][6]) + "</code>. CALCULATED GAIN: <b>" + str(
        round(float(_list[1][11]) + float(_list[1][10]) - float(
            _list[0][11]) - float(_list[0][10]),
              5)) + "€</b>" + "\nSPREAD: <b>" + str(
        round(float(_list[0][7].replace(",", ".")) - float(_list[0][4].replace(",", ".")))) + "€</b>").replace(" ",
                                                                                                               "%20")
    bot_token = tg_data["token"]
    bot_chatID = tg_data["app_id"]
    send_text = 'https://api.telegram.org/bot' + bot_token + '/sendMessage?chat_id=' + bot_chatID + '&parse_mode=HTML&text=' + message
    requests.get(send_text)


def num(num):
    s = str(abs(num))
    if num >= 0:
        sign = "+"
    else:
        sign = "-"
    spc = 7 - len(s)
    s = sign + (" " * spc) + s
    if (int(num) - num) == 0:
        s = s + "0"
    if s[6] == ".":
        s = s + "0"
        s = kill_char(s, 1)
    return s


def kill_char(string, n):
    begin = string[:n]
    end = string[n + 1:]
    return begin + end


def arbomonitor(s, only_see, last_h, name):
    data = s.recv(1024)
    send_json = json.dumps(
        {"timestamp": str(int(datetime.datetime.now(datetime.timezone.utc).timestamp())), "status": not only_see,
         "latency": int(last_h), "name": name})
    byt = send_json.encode()
    s.send(byt)
    return data.decode()


def check_api_connection():
    trt_conn = requests.get("https://www.therocktrading.com/it/")
    if not trt_conn.ok:
        print("HTTP " + str(trt_conn.status_code) + ", error 10")
        return False
    else:
        trt_conn = requests.get("https://api.therocktrading.com/")
        if not trt_conn.ok:
            print("HTTP " + str(trt_conn.status_code) + ", error 12")
            return False
    bnb_conn = requests.get("https://api1.binance.com/")
    if not bnb_conn.ok:
        print("HTTP " + str(bnb_conn.status_code) + ", error 11")
        return False
    else:
        bnb_conn = requests.get("https://api.therocktrading.com/")
        if not bnb_conn.ok:
            print("HTTP " + str(bnb_conn.status_code) + ", error 13")
            return False
        else:
            bnb_conn = requests.get("https://api1.binance.com/wapi/v3/systemStatus.html")
            if bnb_conn.json()["status"] != 0:
                print("HTTP " + str(bnb_conn.status_code) + ", error 14")
                return False
    return True


def getaction(q):
    s_act = socket.socket()
    s_act.bind(("0.0.0.0", 31000))
    s_act.listen(5)
    s_act.settimeout(0.5)
    while True:
        try:
            conn, address = s_act.accept()
            data = conn.recv(1024)
            print(data.decode())
            q.put(data.decode())
            conn.close()
        except socket.timeout as err:
            log("ERR", err)
            pass


def last(q):
    while not q.empty():
        a = q.get()
        if q.empty():
            return a


def auto_balancer(balance_score, op, exchange_list, config, hype):
    price = 0
    # MARKET ORDER
    all_balance = op.balancethreading()
    try:
        if abs(balance_score) > config['balancing']['threshold']:
            if balance_score < 0:
                if not config['balancing']['banking']:
                    depth = (all_balance["bnbbtc"] + all_balance["trtbtc"]) / 2
                    resp_dict = op.tradethreading("sell", "trt", "BTCEUR", depth, price,
                                                  "buy", exchange_list[1], "BTCEUR", depth, price)
                    return resp_dict
                else:
                    op.withdraw(exchange_list[1], 'EUR', config['balancing']['IBAN'],
                                abs(all_balance["bnbeur"] - all_balance["trteur"]))
                    hype.transfer(config['balancing']['IBAN_exch_1'], config['balancing']['recipient_exch_1'],
                                  abs(all_balance["bnbeur"] - all_balance["trteur"]),
                                  config['balancing']['reference_exch_0'])
                    op.withdraw(exchange_list[0], 'BTC', config['balancing']['btc_addr_1'],
                                abs(all_balance["bnbbtc"] - all_balance["trtbtc"]))
            if balance_score > 0:
                if not config['balancing']['banking']:
                    depth = (all_balance["bnbbtc"] + all_balance["trtbtc"]) / 2
                    resp_dict = op.tradethreading("buy", "trt", "BTCEUR", depth, price,
                                                  "sell", exchange_list[1], "BTCEUR", depth, price)
                    return resp_dict
                else:
                    op.withdraw(exchange_list[0], 'EUR', config['balancing']['IBAN'],
                                abs(all_balance["bnbeur"] - all_balance["trteur"]))
                    hype.transfer(config['balancing']['IBAN_exch_1'], config['balancing']['recipient_exch_0'],
                                  abs(all_balance["bnbeur"] - all_balance["trteur"]),
                                  config['balancing']['reference_exch_0'])
                    op.withdraw(exchange_list[1], 'BTC', config['balancing']['btc_addr_1'],
                                abs(all_balance["bnbbtc"] - all_balance["trtbtc"]))
    except Exception as err:
        log("ERR", err)
        return str(err)
    return 0


def info(exchange_a, exchange_b, all_balance, asks, bids, taker_fee, maker_fee=None):
    print(f"[i] ASK %s : %.2f                              EUR %s BAL : {Fore.RED}%.5f{Style.RESET_ALL}" % (
        exchange_b.upper(), asks[exchange_b], exchange_b.upper(), all_balance[exchange_b + "eur"]))
    print(f"[i] BID %s : %.2f                              BTC %s BAL : {Fore.RED}%.8f{Style.RESET_ALL}" % (
        exchange_a.upper(), bids[exchange_a], exchange_a.upper(), all_balance[exchange_a + "btc"]))
    print("[i]                           DIFFERENCE: %s" % (num(round(bids[exchange_a] - asks[exchange_b], 2))))
    print(f"[i]                           DIFF + FEE: {Fore.RED}%s{Style.RESET_ALL}"
          % (num(
        round((bids[exchange_a] * (1 - taker_fee[exchange_a])) - (asks[exchange_b] * (1 + taker_fee[exchange_b])), 2))))
    print(f"[i] ASK %s : %.2f                              EUR %s BAL : {Fore.RED}%.5f" % (
        exchange_a.upper(), asks[exchange_a], exchange_a.upper(), all_balance[exchange_a + "eur"]))
    print(
        f"{Style.RESET_ALL}[i] BID %s : %.2f                              BTC %s BAL : {Fore.RED}%.8f{Style.RESET_ALL}" % (
            exchange_b.upper(), bids[exchange_b], exchange_b.upper(), all_balance[exchange_b + "btc"]))
    print(
        f"[i]                           DIFFERENCE: %s                             TOT EUR: {Fore.GREEN}%.8f{Style.RESET_ALL}" % (
            num(round(bids[exchange_b] - asks[exchange_a], 2)),
            all_balance[exchange_b + "eur"] + all_balance[exchange_a + "eur"]))
    print(
        f"[i]                           DIFF + FEE: {Fore.RED}%s{Style.RESET_ALL}                             TOT BTC: {Fore.GREEN}%.8f{Style.RESET_ALL}                       PF VALUE: {Fore.GREEN}%.4f{Style.RESET_ALL}"
        % (num(
            round((bids[exchange_b] * (1 - taker_fee[exchange_b])) - (asks[exchange_a] * (1 + taker_fee[exchange_a])),
                  2)),
           all_balance[exchange_b + "btc"] + all_balance[exchange_a + "btc"],
           all_balance[exchange_b + "eur"] + all_balance[exchange_a + "eur"] + (
                   (all_balance[exchange_b + "btc"] + all_balance[exchange_a + "btc"]) * bids[exchange_a])))
# TODO: FUNZIONE AUTO-BILANCIAMENTO
# TODO: FUNZIONE PER CONTROLLARE IL SALDO BNB E AGGIUNGERLO QUANDO SERVE
# TODO: SISTEMARE INTEGRAZIONE DATABASE
# TODO: CONTROLLARE IL BOOK PIU IN PROFONDITA PER EVITARE BOOK VUOTI
# TODO: FIX VARI

# TODO: AI PREDICTION ON NEXT PRICE FOR BOTH EXCHANGE, AND/OR PREDICTION OF ARBITRAGE VALUES
