import datetime
import json
import queue
import sys
import threading
import time
from multiprocessing import Process
import signal
import pandas
from colorama import Fore
from colorama import Style
from openpyxl import load_workbook

import data_visual
from trade import Operation

_list = []
_trade_list = []
exchange_list = []
d = {}
login_data = json.loads(open("keydict.txt", "r").readline().strip().replace("\n", ""))
try:
    trt_apikey = str(login_data["trt_apikey"])
    trt_secret = str(login_data["trt_secret"])
    exchange_list.append("trt")
except:
    trt_apikey = ""
    trt_secret = ""
    pass
try:
    krk_apikey = str(login_data["krk_apikey"])
    krk_secret = str(login_data["krk_secret"])
    exchange_list.append("krk")
except:
    krk_apikey = ""
    krk_secret = ""
    pass
try:
    bnb_apikey = str(login_data["bnb_apikey"])
    bnb_secret = str(login_data["bnb_secret"])
    exchange_list.append("bnb")
except:
    bnb_apikey = ""
    bnb_secret = ""
    pass

with open("config.txt") as f:
    for line in f:
        (key, val) = line.replace(" ", "").split("=")
        d[key] = val

bnb_que = queue.Queue()
trt_que = queue.Queue()
all_balance = dict()
time_list = []


def main():
    op = Operation(trt_apikey, trt_secret, krk_apikey, krk_secret, bnb_apikey, bnb_secret, exchange_list)
    op.threadCreation()
    time.sleep(2)
    fee = op.feethreading()
    taker_fee_trt = float(fee["feetrt"]) / 100
    taker_fee_bnb = 75 * float(fee["fee" + exchange_list[1]]) / 100
    print("GOT FEE FROM EXCHANGE; %s: %f;    %s: %f" % (
        exchange_list[0].upper(), taker_fee_trt, exchange_list[1].upper(), taker_fee_bnb))
    checkbalance = True
    bal_list = False
    if d["graph"].lower() == "true":
        g = Process(target=data_visual.ru)
        g.start()
    while 1:
        try:
            eff = 0
            prod = 0
            if checkbalance:
                print(f"{Fore.YELLOW}[#] RETRIEVING BALANCE{Style.RESET_ALL}")
                all_balance = op.balancethreading()
                checkbalance = False
            time.sleep(1 / int(d["rate"]))
            _start_time = time.time()
            _query_time = time.time()
            price_dict = op.querythread()
            _query_time = time.time() - _query_time
            try:
                asks_data_bnb = price_dict["bnb"]['asks'][0]
                bids_data_bnb = price_dict["bnb"]['bids'][0]
                asks_data_trt = price_dict["trt"]['asks'][0]
                bids_data_trt = price_dict["trt"]['bids'][0]
            except TypeError:
                print(f"{Fore.RED}[#] ERROR WHILE FETCHING DATA [typeError - nonetype]{Style.RESET_ALL}")
                continue
            asks_krk = round(float(price_dict["bnb"]['asks'][0][0]), 2)
            bids_krk = round(float(price_dict["bnb"]['bids'][0][0]), 2)
            asks_trt = round(float(price_dict["trt"]['asks'][0]['price']), 2)
            bids_trt = round(float(price_dict["trt"]['bids'][0]['price']), 2)
            if bal_list:
                _trade_list.append(
                    ["", "", "", "", "", "", "", "", all_balance["bnbbtc"], all_balance["trtbtc"],
                     all_balance["bnbeur"],
                     all_balance["trteur"], all_balance["trteur"] + all_balance["bnbeur"] + (
                             all_balance["bnbbtc"] + all_balance["trtbtc"]) * last_bid])
            last_bid = 0
            last_ask = 0
            depth = 0
            bal_list = False
            print(
                f"{Fore.LIGHTCYAN_EX}[i] %s{Style.RESET_ALL}          INDEX: {Fore.LIGHTCYAN_EX}%s{Style.RESET_ALL} " % (
                    datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), str(int(time.time()))[-4:]))

            print(f"[i] ASK %s : %.2f                              EUR %s BAL : {Fore.RED}%.8f{Style.RESET_ALL}" % (
                exchange_list[1].upper(), asks_krk, exchange_list[1].upper(), all_balance["bnbeur"]))
            print(f"[i] BID %s : %.2f                              BTC %s BAL : {Fore.RED}%.8f{Style.RESET_ALL}" % (
                exchange_list[0].upper(), bids_trt, exchange_list[0].upper(), all_balance["trtbtc"]))
            print("[i]                           DIFFERENCE:", round(bids_trt - asks_krk, 2))
            print(f"[i]                           DIFF + FEE: {Fore.RED}%.2f{Style.RESET_ALL}"
                  % (round((bids_trt * (1 - taker_fee_trt)) - (asks_krk * (1 + taker_fee_bnb)), 2)))
            print(f"[i] ASK %s : %.2f                              EUR %s BAL : {Fore.RED}%.8f{Style.RESET_ALL}" % (
                exchange_list[0].upper(), asks_trt, exchange_list[0].upper(), all_balance["trteur"]))
            print(f"[i] BID %s : %.2f                              BTC %s BAL : {Fore.RED}%.8f{Style.RESET_ALL}" % (
                exchange_list[1].upper(), bids_krk, exchange_list[1].upper(), all_balance["bnbbtc"]))
            print(
                f"[i]                           DIFFERENCE: %.2f                             TOT EUR: {Fore.GREEN}%.8f{Style.RESET_ALL}" % (
                    round(bids_krk - asks_trt, 2), all_balance["bnbeur"] + all_balance["trteur"]))
            print(
                f"[i]                           DIFF + FEE: {Fore.RED}%.2f{Style.RESET_ALL}                            TOT BTC: {Fore.GREEN}%.8f{Style.RESET_ALL}                       PF VALUE: {Fore.GREEN}%.4f{Style.RESET_ALL}"
                % (round((bids_krk * (1 - taker_fee_bnb)) - (asks_trt * (1 + taker_fee_trt)), 2),
                   all_balance["bnbbtc"] + all_balance["trtbtc"], all_balance["bnbeur"] + all_balance["trteur"] + (
                           (all_balance["bnbbtc"] + all_balance["trtbtc"]) * bids_trt)))
            print("[i] FETCHED FEE       %s: %.4f%%;      %s: %.4f%%" % (
                exchange_list[0].upper(), taker_fee_trt * 100, exchange_list[1].upper(), taker_fee_bnb * 100))

            if (bids_trt * (1 - taker_fee_trt)) - (asks_krk * (1 + taker_fee_bnb)) > 0:
                low_balance = False
                print(f"{Fore.CYAN}[#] %.2f < %.2f BUY %s | SELL TRT DIFF: %.2f (MENO FEE): %.3f" % (
                    asks_krk, bids_trt, exchange_list[1].upper(), bids_trt - asks_krk,
                    (bids_trt * (1 + taker_fee_trt)) - (asks_krk * (1 + taker_fee_bnb))))
                depth = min(bids_data_trt['amount'],
                            float(asks_data_bnb[1]))
                balance = min(all_balance["trtbtc"], all_balance["bnbeur"] / asks_krk)
                if balance < depth:
                    depth = balance
                    print(f"{Fore.MAGENTA}[#] PARTIAL FILLING, BALANCE LOWER THAN DEPTH{Style.RESET_ALL}")
                    if depth == 0:
                        print(f"{Fore.RED}[#] BALANCE IS LOW, PLEASE DEPOSIT TO CONTINUE{Style.RESET_ALL}")
                        low_balance = True
                if not low_balance and (asks_krk * depth) > int(d["min_balance"]):
                    print(f"{Fore.CYAN}[#] DEPTH %f BTC" % depth)
                    eff = (depth * bids_trt * (1 - taker_fee_trt)) - (depth * asks_krk * (1 + taker_fee_bnb))
                    prod = eff / (depth * bids_trt)
                    print("[#] GAIN DOPO FEE EFF %f € | PROD %f ¢/€" % (eff, prod * 100))
                    print(f"[#] NEED %.3f EUR | %f BTC{Style.RESET_ALL}" % (asks_krk * depth, depth))
                    last_ask = asks_krk
                    last_bid = bids_trt
                    if prod * 100 > float(d["prod_threshold"]):
                        checkbalance = True
                        print(f"{Fore.GREEN}[#] TRADE{Style.RESET_ALL}")
                        resp_dict = op.tradethreading("sell", "trt", "BTCEUR", depth, last_bid, "buy", "bnb", "BTCEUR",
                                                      depth,
                                                      last_ask)

                        if resp_dict["bnb"] == "ERROR" or resp_dict["trt"][1] == "ERROR":
                            print(f"{Fore.RED}[$] TRADE ERROR MSG: [%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"][0].upper(), resp_dict["bnb"]))
                            op.cancelthreading()
                        else:
                            print(f"{Fore.GREEN}[#] SOUNDS GOOD! ORDER STATUS:[%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"].upper(), resp_dict["bnb"]))
                            bal_list = True
                            time.sleep(int(d["sleep_check_order"]))
                            _trade_list.append(
                                [datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "buy", exchange_list[1], depth,
                                 last_ask,
                                 "sell", "trt", last_bid, all_balance["bnbbtc"], all_balance["trtbtc"],
                                 all_balance["bnbeur"], all_balance["trteur"],
                                 all_balance["trteur"] + all_balance["bnbeur"] + (
                                         all_balance["bnbbtc"] + all_balance["trtbtc"]) * last_bid])
                            op.cancelthreading()
                            # EXECUTED OR SUCCESS
                            # IF BOTH ORDER ARE NOT COMPLETED, DELETE ORDER
                else:
                    print(f"{Fore.RED}[$] TOO LOW BALANCE, PLEASE DEPOSIT{Style.RESET_ALL}")
                    checkbalance = True

            elif (bids_krk * (1 - taker_fee_bnb)) - (asks_trt * (1 + taker_fee_trt)) > 0:
                low_balance = False
                print(f"{Fore.CYAN}[!] %.2f < %.2f BUY TRT | SELL %s DIFF: %.2f (MENO FEE): %.3f{Style.RESET_ALL}" % (
                    asks_trt, bids_krk, exchange_list[1].upper(), bids_krk - asks_trt,
                    (bids_krk * (1 + taker_fee_bnb)) - (asks_trt * (1 + taker_fee_trt))))
                depth = min(asks_data_trt['amount'],
                            float(bids_data_bnb[1]))
                balance = min(all_balance["bnbbtc"], all_balance["trteur"] / asks_trt)
                if balance < depth:
                    depth = balance
                    print(f"{Fore.MAGENTA}[#] PARTIAL FILLING, BALANCE LOWER THAN DEPTH{Style.RESET_ALL}")
                    if depth == 0:
                        print(f"{Fore.MAGENTA}[#] BALANCE IS LOW, PLEASE DEPOSIT TO CONTINUE{Style.RESET_ALL}")
                        low_balance = True
                else:
                    print(f"{Fore.GREEN}[#] COMPLETE FILLING{Style.RESET_ALL}")
                if not low_balance and (asks_krk * depth) > int(d["min_balance"]):
                    print(f"{Fore.CYAN}[!] DEPTH %f BTC" % depth)
                    eff = (depth * bids_krk * (1 - taker_fee_bnb)) - (depth * asks_trt * (1 + taker_fee_trt))
                    prod = eff / (depth * bids_krk)
                    print("[i] GAIN DOPO FEE EFF %f € | PROD %f ¢/€" % (eff, prod * 100))
                    print(f"[i] NEED %.3f EUR | %f BTC{Style.RESET_ALL}" % (asks_trt * depth, depth))
                    last_ask = asks_trt
                    last_bid = bids_krk
                    if prod * 100 > float(d["prod_threshold"]):
                        checkbalance = True
                        print(f"{Fore.GREEN}[#] TRADE{Style.RESET_ALL}")
                        resp_dict = op.tradethreading("buy", "trt", "BTCEUR", depth, last_ask,
                                                      "sell", exchange_list[1], "BTCEUR", depth, last_bid)
                        if resp_dict['bnb'] == "ERROR" or resp_dict['bnb'] == "NEW" or resp_dict[
                            'trt'].upper() == "ERROR":
                            print(f"{Fore.RED}[$] TRADE ERROR MSG: [%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"][0].upper(), resp_dict["bnb"]))
                            op.cancelthreading()
                        else:
                            print(f"{Fore.GREEN}[#] SOUNDS GOOD! ORDER NO:[%s, %s]{Style.RESET_ALL}" % (
                                resp_dict["trt"].upper(), resp_dict["bnb"]))
                            bal_list = True
                            time.sleep(int(d["sleep_check_order"]))
                            _trade_list.append(
                                [datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "BUY", "TRT",
                                 str(depth).replace(".", ","), str(last_ask).replace(".", ","),
                                 "SELL", exchange_list[1].upper(), str(last_bid).replace(".", ","),
                                 all_balance["bnbbtc"],
                                 all_balance["trtbtc"],
                                 all_balance["bnbeur"], all_balance["trteur"],
                                 all_balance["trteur"] + all_balance["bnbeur"] + (
                                         all_balance["bnbbtc"] + all_balance["trtbtc"]) * last_bid])
                            op.cancelthreading()
                else:
                    print(f"{Fore.RED}[$] TOO LOW BALANCE, PLEASE DEPOSIT{Style.RESET_ALL}")
                    checkbalance = True
            _end_time = time.time()
            totaltime = _end_time - _start_time
            time_list.append(int(totaltime * 1000))
            _list.append([datetime.datetime.now(), asks_krk, bids_krk, asks_trt, bids_trt,
                          all_balance["bnbbtc"], all_balance["trtbtc"], all_balance["bnbeur"],
                          all_balance["trteur"], eff, prod, int(totaltime * 1000), round(bids_trt - asks_krk, 2),
                          round((bids_trt * (1 - taker_fee_trt)) - (asks_krk * (1 + taker_fee_bnb)), 2),
                          round(bids_krk - asks_trt, 2),
                          round((bids_krk * (1 - taker_fee_bnb)) - (asks_trt * (1 + taker_fee_trt)), 2)])
            if int(_end_time % int(d["save_interval"])) == 0:
                print(f"{Fore.YELLOW}[!] SAVING...{Style.RESET_ALL}")
                if _list:
                    save_data_thread = threading.Thread(target=save_data, args=(_list,))
                    save_data_thread.start()

            if int(_end_time % int(d["balance_interval"])) == 0:
                checkbalance = True
            if int(_end_time % int(d["save_trade_interval"])) == 0:
                if _trade_list:
                    print(f"{Fore.YELLOW}[!] SAVING TRADE LIST...{Style.RESET_ALL}")
                    save_trade_thread = threading.Thread(target=save_trade, args=(_trade_list,))
                    save_trade_thread.start()
            if int(_end_time % int(d["fee_interval"])) == 0:
                print(f"{Fore.YELLOW}[!] FETCHING FEE DATA...{Style.RESET_ALL}")
                fee = op.feethreading()
                taker_fee_trt = float(fee["fee" + exchange_list[0]]) / 100
                taker_fee_krk = float(fee["fee" + exchange_list[1]]) / 100

            print(
                f"[-] ------------------------------------------------- {Fore.YELLOW}%d ms{Style.RESET_ALL} (%d ms(q) + %d ms(p)) - avg last %d ({Fore.YELLOW}%d ms{Style.RESET_ALL}) - global avg ({Fore.YELLOW}%d ms{Style.RESET_ALL})" % (
                    int(totaltime * 1000), int(_query_time * 1000),
                    (int(totaltime * 1000) - int(_query_time * 1000)), min(100, len(time_list)),
                    sum(time_list[-100:]) / min(100, len(time_list)), sum(time_list) / len(time_list)))
        except KeyboardInterrupt:
            sys.exit()


def save_data(_list):
    df = pandas.DataFrame(_list)
    try:
        df.to_csv('filev2.csv', index=False, sep=';', mode='a', header=False, decimal=',')
        # append(df, filename='filev2.xlsx', startrow=None, sheet_name='Sheet1', truncate_sheet=True,engine="openpyxl")
        _list.clear()
    except FileNotFoundError:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATALOG [file not found]{Style.RESET_ALL}")
    except PermissionError:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATALOG [resource busy]{Style.RESET_ALL}")
    except:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATALOG [generic error]{Style.RESET_ALL}")


def save_trade(_list):
    df = pandas.DataFrame(_list)
    try:
        # with open('file_trade.xlsx', 'a') as f:
        df.to_csv("file_trade.csv", sep=';', mode='a', index=False, header=False, decimal=',')
        # append(df, filename='file_trade.xlsx', startrow=None, sheet_name='Sheet1', truncate_sheet=True,engine="openpyxl")
        _list.clear()
    except FileNotFoundError:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO TRADELIST [file not found]{Style.RESET_ALL}")
    except PermissionError:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO TRADELIST [resource busy]{Style.RESET_ALL}")
    except TypeError:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO TRADELIST [type error]{Style.RESET_ALL}")
    except:
        print(f"{Fore.RED}[ERR] ERRORE SALVATAGGIO DATALOG [generic error]{Style.RESET_ALL}")


def append(df, filename, startrow=None, sheet_name='Sheet1', truncate_sheet=True, engine="xlrd"):
    writer = pandas.ExcelWriter(filename, engine=engine)
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

        # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow)

    # save the workbook
    writer.save()

def exit_gracefully(self):
    self.kill_now = True
if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
